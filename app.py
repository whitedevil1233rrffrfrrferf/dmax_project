from flask import Flask,render_template,request,redirect,url_for,jsonify,flash,session
from dotenv import load_dotenv
import os
from openpyxl import load_workbook
from flask_sqlalchemy import SQLAlchemy
from googleapiclient.discovery import build
from google.oauth2 import service_account
import json
from datetime import datetime
from sqlalchemy import extract, func, or_
import requests
from google_auth_oauthlib.flow import Flow
from google.oauth2.credentials import Credentials
os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '0'
app=Flask(__name__)
CLIENT_ID = os.environ.get('CLIENT_ID')
CLIENT_SECRET = os.environ.get('CLIENT_SECRET')
REDIRECT_URI='http://127.0.0.1:5000/call_back'
SCOPES = ['openid', 'https://www.googleapis.com/auth/userinfo.email', 'https://www.googleapis.com/auth/userinfo.profile']


flow = Flow.from_client_config(
    {
        "web": {
            "client_id": CLIENT_ID,
            "client_secret": CLIENT_SECRET,
            "redirect_uris": [REDIRECT_URI],
            "auth_uri": "https://accounts.google.com/o/oauth2/auth",
            "token_uri": "https://oauth2.googleapis.com/token",
        }
    },
    scopes=SCOPES
)

excel_path = os.path.join(app.static_folder,'files','DMAX-2024-Live.xlsx')
workbook = load_workbook(excel_path)
sheet = workbook.active
def find_next_available_row(sheet):
    for row in range(1, sheet.max_row + 1):
        if all([cell.value in [None, ""] for cell in sheet[row]]):
            
            return row
        
    return sheet.max_row + 1 

def credentials_to_dict(credentials):
    return {
        'token': credentials.token,
        'refresh_token': credentials.refresh_token,
        'token_uri': credentials.token_uri,
        'client_id': credentials.client_id,
        'client_secret': credentials.client_secret,
        'scopes': credentials.scopes
    }

app.config['SECRET_KEY'] = 'your_secret_key'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///employees.db'
app.config['SQLALCHEMY_BINDS']={
    'dform':'sqlite:///dform.db',
    'emp_info':'sqlite:///empinfo.db',
    'op_excellence':'sqlite:///opexcellence.db'
}
db = SQLAlchemy(app)


# helper functions -------





class Dform(db.Model):
    __tablename__ = 'login'
    __bind_key__="dform"
    id = db.Column(db.Integer, primary_key=True)
    employee_name=db.Column(db.String(100),nullable=False)
    employee_id=db.Column(db.String(100),nullable=False)
    employee_email=db.Column(db.String(100),nullable=False)
    today_date=db.Column(db.String(100),nullable=False)
    project=db.Column(db.String(100),nullable=False)
    designation=db.Column(db.String(100),nullable=False)
    test_case_creation_target= db.Column(db.Integer)
    test_case_creation_actual=db.Column(db.Integer)
    test_case_updation_target=db.Column(db.Integer)
    test_case_updation_actual=db.Column(db.Integer)
    test_case_execution_target=db.Column(db.Integer)
    test_case_execution_actual=db.Column(db.Integer)
    defects_found_target=db.Column(db.Integer)
    defects_found_actual=db.Column(db.Integer)
    test_scripts_creation_target=db.Column(db.Integer)
    test_scripts_creation_actual=db.Column(db.Integer)
    test_scripts_updation_target=db.Column(db.Integer)
    test_scripts_updation_actual=db.Column(db.Integer)
    test_scripts_execution_target=db.Column(db.Integer)
    test_scripts_execution_actual=db.Column(db.Integer)
    site_Scrub_target=db.Column(db.Integer)
    site_Scrub_actual=db.Column(db.Integer)
    project_doc_target=db.Column(db.Integer)
    project_doc_actual=db.Column(db.Integer)
    internal_Review_target=db.Column(db.Integer)
    internal_Review_actual=db.Column(db.Integer)
    regression_cycle_target=db.Column(db.Integer)
    regression_cycle_actual=db.Column(db.Integer)
    req_anal_target=db.Column(db.Integer)
    req_anal_actual=db.Column(db.Integer)
    end_cases_exec_target=db.Column(db.Integer)
    end_cases_exec_actual=db.Column(db.Integer)
    task_coverage_score_target=db.Column(db.Integer)
    task_coverage_score_actual=db.Column(db.Integer)
    assessment_score_target=db.Column(db.Integer)
    assessment_score_actual=db.Column(db.Integer)
    assessment_re_score_target=db.Column(db.Integer)
    assessment_re_score_actual=db.Column(db.Integer)
    cert_score_target=db.Column(db.Integer)
    cert_score_actual=db.Column(db.Integer)
    cert_re_score_target=db.Column(db.Integer)
    cert_re_score_actual=db.Column(db.Integer)
    new_features_imp_target=db.Column(db.Integer)
    new_features_imp_actual=db.Column(db.Integer)
    defects_fixed_target=db.Column(db.Integer)
    defects_fixed_actual=db.Column(db.Integer)
    enhancements_target=db.Column(db.Integer)
    enhancements_actual=db.Column(db.Integer)
    fig_desgns_target=db.Column(db.Integer)
    fig_desgns_actual=db.Column(db.Integer)
    doc_update_target=db.Column(db.Integer)
    doc_update_actual=db.Column(db.Integer)
    research_target=db.Column(db.Integer)
    research_actual=db.Column(db.Integer)
    inv_defs=db.Column(db.Integer)
    spel_errors=db.Column(db.Float)
    client_esc=db.Column(db.Integer)
    tst_cases_missing=db.Column(db.Integer)
    att=db.Column(db.Integer)
    dtouch=db.Column(db.Integer)
    new_init=db.Column(db.Integer)
    defects_verification_target=db.Column(db.Integer)
    defects_verification_actual=db.Column(db.Integer)
    target=db.Column(db.Integer)
    actual=db.Column(db.Integer)
    production=db.Column(db.Integer)
    quality=db.Column(db.Integer)
    attendance=db.Column(db.Integer)
    skill=db.Column(db.Integer)
    new_initiatives=db.Column(db.Integer)
    Dmax_score=db.Column(db.Integer)
    

# Employee Model
class Employee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    emp_id = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(100), nullable=False)
    email= db.Column(db.String(100), nullable=False)
    name= db.Column(db.String(100), nullable=False)
class Employee_information(db.Model):
    __bind_key__="emp_info"
    id = db.Column(db.Integer, primary_key=True)
    emp_name = db.Column(db.String(100),nullable=False)
    emp_id = db.Column(db.String(100),nullable=False)
    emp_email = db.Column(db.String(100),nullable=False)
    emp_date = db.Column(db.String(100),nullable=False)
    emp_project = db.Column(db.String(100),nullable=False)
    emp_designation = db.Column(db.String(100),nullable=False)
    test_case_creation_target= db.Column(db.Integer)
    test_case_updation_target=db.Column(db.Integer)
    test_case_execution_target=db.Column(db.Integer)
    defects_found_target=db.Column(db.Integer)
    test_scripts_creation_target=db.Column(db.Integer)
    test_scripts_updation_target=db.Column(db.Integer)
    test_scripts_execution_target=db.Column(db.Integer)
    site_Scrub_target=db.Column(db.Integer)
    project_doc_target=db.Column(db.Integer)
    internal_Review_target=db.Column(db.Integer)
    regression_cycle_target=db.Column(db.Integer)
    req_anal_target=db.Column(db.Integer)
    end_cases_exec_target=db.Column(db.Integer)
    task_coverage_score_target=db.Column(db.Integer)
    assessment_score_target=db.Column(db.Integer)
    assessment_re_score_target=db.Column(db.Integer)
    cert_score_target=db.Column(db.Integer)
    cert_re_score_target=db.Column(db.Integer)
    new_features_imp_target=db.Column(db.Integer)
    defects_fixed_target=db.Column(db.Integer)
    enhancements_target=db.Column(db.Integer)
    fig_desgns_target=db.Column(db.Integer)
    doc_update_target=db.Column(db.Integer)
    research_target=db.Column(db.Integer)
    inv_defs=db.Column(db.Integer)
    spel_errors=db.Column(db.Float)
    client_esc=db.Column(db.Integer)
    tst_cases_missing=db.Column(db.Integer)
    att=db.Column(db.Integer)
    dtouch=db.Column(db.Integer)
    new_init=db.Column(db.Integer)
    defects_verification_target=db.Column(db.Integer)
    reporting_manager=db.Column(db.String(100))

class OperationalExcellence(db.Model):
    __bind_key__="op_excellence"
    
    id = db.Column(db.Integer, primary_key=True)
    emp_id = db.Column(db.String(100), nullable=False, unique=True)
    attendance_score = db.Column(db.Float ,default=0.0)
    dtouch_score = db.Column(db.Float,default=0.0)
    new_init_score = db.Column(db.Float,default=0.0)


def get_logged_in_user_details():
    """
    Retrieves the logged-in user's name from the Employee table based on session data.
    """
    # Check if the user logged in with username/password
    if 'username' in session:
        username = session['username']
        user = Employee.query.filter_by(emp_id=username).first()  # Match emp_id with the username
        if user:
            return {"name": user.name, "role": user.role,"email":user.email}

    # Check if the user logged in with Google Sign-In (using email)
    if 'email' in session:
        email = session['email']
        user = Employee.query.filter_by(email=email).first()  # Match email with the logged-in user's email
        if user:
            return {"name": user.name, "role": user.role}

    # If no user is found, return None or an appropriate message
    return None

def get_filtered_employees(base_query, search_query, selected_month, selected_date):
    if search_query:
        base_query = base_query.filter(func.lower(Dform.employee_name) == search_query)
    if selected_month:
        base_query = base_query.filter(extract('month', Dform.today_date) == int(selected_month))
    if selected_date:
        base_query = base_query.filter_by(today_date=selected_date)
    return base_query.all()

@app.context_processor
def custom_global_variable():
    role = None

    # Check if the user is logged in (email is in session)
    if 'email' in session:
        user_email = session['email']

        # Query the user role from the database
        user = Employee.query.filter_by(email=user_email).first()
        if user:
            role = user.role  # Assuming `role` is a column in your User model
            print("role",role)

    if 'username' in session :
        username = session['username']
        user = Employee.query.filter_by(emp_id=username).first()
        if user:
            role=user.role
            print("role",user)         
    # Return the role to all templates as a global variable
    return {'user_role': role}

@app.template_filter('get_attr')
def get_attr(obj, attr):
    """Fetches an attribute from an object safely."""
    return getattr(obj, attr, 'N/A')

@app.route('/form',methods=["GET","POST"])
def home():
    if 'username' in session :
        username = session['username']
        employee = Employee.query.filter_by(emp_id=username).first()
        if employee:
            role = employee.role
            print(role)
    elif 'email' in session:
        email=session['email']
        employee=Employee.query.filter_by(email=email).first()
        if employee:
            role=employee.role
            print(role)      
            
    else:
       
        return redirect(url_for('sign'))
        

    if request.method=="POST":
        
        workbook = load_workbook(excel_path)
        sheet = workbook.active
        
        next_row = find_next_available_row(sheet)
        
        field_to_column = {
            "employee_name": 'A',
            "employee_id": 'B',
            "employee_email": 'C',
            "today_date": 'D',
            "project": 'E',
            "designation": 'F',
            "test_case_creation_target": 'G',
            "test_case_creation_actual": 'H',
            "test_case_updation_target": 'I',
            "test_case_updation_actual": 'J',
            "test_case_execution_target": 'K',
            "test_case_execution_actual": 'L',
            "defects_found_target":'M',
            "defects_found_actual":'N',
            "defects_verification_target":'O',
            "defects_verification_actual":'P',
            "test_scripts_creation_target":'Q',
            "test_scripts_creation_actual":'R',
            "test_scripts_updation_target":'S',
            "test_scripts_updation_actual":'T',
            "test_scripts_execution_target":'U',
            "test_scripts_execution_actual":'V',
            "site_Scrub_target":'AG',
            "site_Scrub_actual":'AH',
            "project_doc_target":'W',
            "project_doc_actual":'X',
            "internal_Review_target":'Y',
            "internal_Review_actual":'Z',
            "regression_cycle_target":'AA',
            "regression_cycle_actual":'AB',
            "req_anal_target":'AC',
            "req_anal_actual":'AD',
            "end_cases_exec_target":'AE',
            "end_cases_exec_actual":'AF',
            "task_coverage_score_target":'AI',
            "task_coverage_score_actual":'AJ',
            "assessment_score_target":'AK',
            "assessment_score_actual":'AL',
            "assessment_re_score_target":'AM',
            "assessment_re_score_actual":'AN',
            "cert_score_target":"AO",
            "cert_score_actual":'AP',
            "cert_re_score_target":'AQ',
            "cert_re_score_actual":'AR',
            "new_features_imp_target":'AS',
            "new_features_imp_actual":'AT',
            "defects_fixed_target":'AU',
            "defects_fixed_actual":'AV',
            "enhancements_target":'AW',
            "enhancements_actual":'AX',
            "fig_desgns_target":'AY',
            "fig_desgns_actual":'AZ',
            "doc_update_target":'BA',
            "doc_update_actual":'BB',
            "research_target":'BC',
            "research_actual":'BD',
            "inv_defs":'BE',
            "spel_errors":'BF',
            "client_esc":'BG',
            "tst_cases_missing":'BH',
            "att":'BI',
            "dtouch":'BJ',
            "new_init":'BK',    
        }
        form_data = {}
        row_values = []
        for field, column in field_to_column.items():
            value = request.form.get(field)
            value = value.strip() if value else ''
            if value and value.replace('.', '', 1).isdigit():
                value = float(value)
            form_data[field] = value
            # sheet[f'{column}{next_row}'] = value
            # row_values.append(value)
        # dictionary for mapping actual to target    
        actual_to_target_mapping = {}

        for key in field_to_column.keys():
            if key.endswith('_actual'):
                target_key = key.replace('_actual', '_target')  # Replace '_actual' with '_target'
                if target_key in field_to_column:  # Check if target_key exists
                    actual_to_target_mapping[key] = target_key   

        results = {}

        # Initialize the 'BL' sum as 0
        operational_excellence=OperationalExcellence.query.filter_by(emp_id=form_data["employee_id"]).first()
        if operational_excellence:
            results['BP']=operational_excellence.attendance_score
            results['BQ'] = operational_excellence.dtouch_score
            results['BR'] = operational_excellence.new_init_score
        else:
            results['BP']=0   
            results['BQ'] = 0
            results['BR'] = 0 
        print(results['BP'])     
        results['BL'] = 0
        results['BM'] = 0
        # Loop through the actual-to-target mapping and apply the formula
        for actual_field, target_field in actual_to_target_mapping.items():
            if form_data[actual_field] > 0:
                results['BL'] +=int(form_data[target_field])
                results['BM'] += int(form_data[actual_field])  
        if results['BM'] != 0 and results['BL'] != 0:  # Check if both BM and BL are not zero
            results['BN'] = ((results['BM'] / results['BL']) * 40 / 100) * 100 
        else:
            results['BN'] = 0     
        if form_data['client_esc'] == 1:  # Check if BG (Client Escalations) is 1
            results['BO'] = 0  # Set BO to 0 if BG is 1
        else:
            sum_invalid_defects_to_test_cases = (
                form_data['inv_defs'] +  # BE: Invalid Defects
                form_data['spel_errors'] +  # BF: Spelling Errors
                form_data['client_esc'] +  # BG: Client Escalations
                form_data['tst_cases_missing']  # BH: Test Cases Missing
            )       
            results['BO'] = ((100 - sum_invalid_defects_to_test_cases) * 0.4 / 100) * 100
            # results['BP'] = int((form_data['att'] * 1 * 10 / 100) * 100) 
            # results['BP']=0
            # results['BQ'] = int(((form_data['dtouch'] * 10 / 100 / 100) * 100)*100)
            # results['BQ'] = 0
            # results['BR'] =int(((form_data['new_init'] * 10 / 100 / 100) * 100)*100)  
            # results['BR'] = 0   
            results['BS'] = sum(
                                results[key] for key in [ 'BN', 'BO', 'BP', 'BQ', 'BR']
                            )
        
        

        new_entry = Dform(
            employee_name=form_data['employee_name'],
            employee_id=form_data['employee_id'],
            employee_email=form_data['employee_email'],
            today_date=form_data['today_date'],
            project=form_data['project'],
            designation=form_data['designation'],
            test_case_creation_target=form_data.get('test_case_creation_target'),
            test_case_creation_actual=form_data.get('test_case_creation_actual'),
            test_case_updation_target=form_data.get('test_case_updation_target'),
            test_case_updation_actual=form_data.get('test_case_updation_actual'),
            test_case_execution_target=form_data.get('test_case_execution_target'),
            test_case_execution_actual=form_data.get('test_case_execution_actual'),
            defects_found_target=form_data.get('defects_found_target'),
            defects_found_actual=form_data.get('defects_found_actual'),
            test_scripts_creation_target=form_data.get('test_scripts_creation_target'),
            test_scripts_creation_actual=form_data.get('test_scripts_creation_actual'),
            test_scripts_updation_target=form_data.get('test_scripts_updation_target'),
            test_scripts_updation_actual=form_data.get('test_scripts_updation_actual'),
            test_scripts_execution_target=form_data.get('test_scripts_execution_target'),
            test_scripts_execution_actual=form_data.get('test_scripts_execution_actual'),
            site_Scrub_target=form_data.get('site_Scrub_target'),
            site_Scrub_actual=form_data.get('site_Scrub_actual'),
            project_doc_target=form_data.get('project_doc_target'),
            project_doc_actual=form_data.get('project_doc_actual'),
            internal_Review_target=form_data.get('internal_Review_target'),
            internal_Review_actual=form_data.get('internal_Review_actual'),
            regression_cycle_target=form_data.get('regression_cycle_target'),
            regression_cycle_actual=form_data.get('regression_cycle_actual'),
            req_anal_target=form_data.get('req_anal_target'),
            req_anal_actual=form_data.get('req_anal_actual'),
            end_cases_exec_target=form_data.get('end_cases_exec_target'),
            end_cases_exec_actual=form_data.get('end_cases_exec_actual'),
            task_coverage_score_target=form_data.get('task_coverage_score_target'),
            task_coverage_score_actual=form_data.get('task_coverage_score_actual'),
            assessment_score_target=form_data.get('assessment_score_target'),
            assessment_score_actual=form_data.get('assessment_score_actual'),
            assessment_re_score_target=form_data.get('assessment_re_score_target'),
            assessment_re_score_actual=form_data.get('assessment_re_score_actual'),
            cert_score_target=form_data.get('cert_score_target'),
            cert_score_actual=form_data.get('cert_score_actual'),
            cert_re_score_target=form_data.get('cert_re_score_target'),
            cert_re_score_actual=form_data.get('cert_re_score_actual'),
            new_features_imp_target=form_data.get('new_features_imp_target'),
            new_features_imp_actual=form_data.get('new_features_imp_actual'),
            defects_fixed_target=form_data.get('defects_fixed_target'),
            defects_fixed_actual=form_data.get('defects_fixed_actual'),
            defects_verification_target=form_data.get('defects_verification_target'),
            defects_verification_actual=form_data.get('defects_verification_actual'),
            enhancements_target=form_data.get('enhancements_target'),
            enhancements_actual=form_data.get('enhancements_actual'),
            fig_desgns_target=form_data.get('fig_desgns_target'),
            fig_desgns_actual=form_data.get('fig_desgns_actual'),
            doc_update_target=form_data.get('doc_update_target'),
            doc_update_actual=form_data.get('doc_update_actual'),
            research_target=form_data.get('research_target'),
            research_actual=form_data.get('research_actual'),
            inv_defs=form_data.get('inv_defs'),
            spel_errors=form_data.get('spel_errors'),
            client_esc=form_data.get('client_esc'),
            tst_cases_missing=form_data.get('tst_cases_missing'),
            att=form_data.get('att'),
            dtouch=form_data.get('dtouch'),
            new_init=form_data.get('new_init'),
            target=results['BL'],
            actual=results['BM'],
            production=results['BN'],
            quality=results['BO'],
            
            # Attendance (BO) and Skill (BP)
            attendance=results['BP'],
            skill=results['BQ'],
            
            # New Initiatives (BQ) and Dmax Score (BS)
            new_initiatives=results['BR'],
            Dmax_score=results['BS'],
        )
        
        # Add to DB and commit the session
        db.session.add(new_entry)
        db.session.commit()
        
        
        return redirect(url_for('home'))
    return render_template('index.html',role=role)

  
    
@app.route('/',methods=["GET","POST"])
def sign():
    session.clear()
    if request.method=="POST":
        username = request.form["username"]  
        password = request.form["password"]
        employee = Employee.query.filter_by(emp_id=username).first()
        if employee and employee.password == password:
            session['username'] = username
            role=employee.role
            if role=="super_admin":
                return redirect(url_for('dmax_table'))
            if role=="admin":
                return redirect(url_for('dmax_table'))
            if role=="manager":
                return redirect(url_for('home'))
            if role=="crewmate":
                return redirect(url_for('view_dscore'))

        else:
            flash("Invalid credentials. Please try again!", "danger")
        
        
    return render_template('sign.html')     

@app.route('/view_dmax')
def view_dmax():
    return "hello"
@app.route('/login')
def login():
    flow.redirect_uri = REDIRECT_URI  
    authorization_url, state = flow.authorization_url()
    session['state'] = state
    
    return redirect(authorization_url) 

@app.route('/call_back')
def google_sign_in():

    flow.fetch_token(authorization_response=request.url)
    
    # Store the credentials in the session
    credentials = flow.credentials
    
    session['credentials'] = credentials_to_dict(credentials)
    
    credentials = Credentials.from_authorized_user_info(session['credentials'])
    response = requests.get('https://www.googleapis.com/oauth2/v3/userinfo', headers={'Authorization': f'Bearer {credentials.token}'})
    if response.status_code == 200:
        user_info = response.json()
        
        user_email = user_info.get('email')  # Extract email from the response
        employee = Employee.query.filter_by(email=user_email).first()
        if not employee:
            flash("Invalid credentials. Please try again!", "danger")
            return redirect(url_for('sign'))
        # Store the email in the session or perform any other actions
        session['email'] = user_email
        if employee.role == 'crewmate':
            return redirect(url_for('view_dscore'))
        if employee.role=='super_admin' or employee.role=="admin":
            return redirect(url_for('dmax_table'))
        
    else:
        print("Failed to fetch user info")
        
    return redirect("/form")
    
@app.route('/read_excel')
def read_excel():
    # Construct the path to the Excel file in the static folder
    

    # Load the Excel workbook
    
    
    # Select the active sheet
    sheet = workbook.active
    
    # Example: Reading data from the first row, first column (A1)
    first_cell_value = sheet['A3'].value

    # Optionally: Process the data further and return it to the template
    return f"Value in A1: {first_cell_value}"


@app.route('/search', methods=['POST'])
def search_employee():
    data = request.json
    employee_name = data.get('employee_name')
    logged_in_user = get_logged_in_user_details()
    logged_in_user_role = logged_in_user.get("role")
    logged_in_user_name = logged_in_user.get("name")
    
    employees_list = []
    # Query the Login table where employee_name matches (case-insensitive search)
    matched_employees = Employee_information.query.filter(Employee_information.emp_name.ilike(f'%{employee_name}%')).all() 
    for emp in matched_employees:
        
    # Create a list of dictionaries containing employee details
        if emp.reporting_manager == logged_in_user_name:
            # Include employee details in the response if the employee's reporting manager matches
            employee_details = {column.name: getattr(emp, column.name) for column in Employee_information.__table__.columns}
            employees_list.append(employee_details)
    
    return jsonify({"employees": employees_list})

@app.route('/register', methods=['GET','POST'])
def register():
    if request.method=="POST":
        username=request.form["username"]
        password=request.form["password"]
        role=request.form["role"]
        email=request.form["email"]
        name=request.form["name"]
        existing_employee = Employee.query.filter(
            or_(
                Employee.emp_id == username,
                Employee.email == email,
                Employee.name == name
            )
        ).first()
        if existing_employee:
            if existing_employee.emp_id == username:
                flash('Employee with that ID already exists', "danger")
            elif existing_employee.email == email:
                flash('Employee with that email already exists', "danger")
            elif existing_employee.name == name:
                flash('Employee with that name already exists', "danger")
            return redirect(url_for('register'))
        new_employee = Employee(emp_id=username, password=password, role=role,email=email,name=name)
        db.session.add(new_employee)
        db.session.commit()
        flash('Employee registered successfully!', 'success')
        
        return redirect(url_for('sign'))
    return render_template("register.html")

@app.route("/no-acess")
def no_access():
    return render_template("no_acess.html")

@app.route("/logout")
def logout():
    session.pop('username',None)
    return redirect(url_for('sign'))

@app.route("/employee_info",methods=["GET","POST"])
def employee_info():
        

    employees = Employee_information.query.all()

    return render_template("employee_info.html",employees=employees)

@app.route('/edit_employee/<int:employee_id>', methods=['GET', 'POST'])
def edit_employee(employee_id):
    # Retrieve employee data from the database based on employee_id
    
    employee =Employee_information.query.get(employee_id)
    projects=['Akyrian','Auxo','Avanti','Bench','Fora Travels','Indihood','IPS','IQHive','LevelBlue','Web Development','Opus Clip','Training']
    designations=['Intern','Jr.QA Engineer','QA Engineer','Sr.QA Engineer','QA Lead']
    if request.method == 'POST':
        
        # Update the employee data with form values
        employee.emp_name = request.form['emp_name']
        employee.emp_id=request.form['emp_id']
        employee.emp_email = request.form['emp_email']
        employee.emp_project = request.form['emp_project']
        employee.emp_designation = request.form['emp_designation']
        employee.reporting_manager=request.form['rep_manager']
        # Save the updated data back to the database
        
        db.session.commit()
        
        return redirect(url_for('employee_info'))
        
    # Render the edit form with existing values
    return render_template('edit_employee.html', employee=employee,projects=projects,designations=designations)



@app.route("/employee_upload" ,methods=['GET', 'POST'])
def employee_upload():
    if request.method=="POST":
        try:
            file = request.files['file']  # Get the uploaded file

            # Load the workbook directly from the file object
            wb = load_workbook(file)  # No need for BytesIO here
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            column_mapping = {
            "Employee Name":"emp_name",	
            "Employee ID":	"emp_id",
            "Employee Email":"emp_email", 
            "Today's Date":"emp_date",	
            "Select your Project":"emp_project",	
            "Designation":"emp_designation",	 
            "Testcase-Creation Target":"test_case_creation_target",
            "Testcase-Updation Target":	"test_case_updation_target",
            "Testcase-Execution Target":"test_case_execution_target",	
            "Defects Found Target":	"defects_found_target",
            "Issue Verification Target":"defects_verification_target",	
            "Test-scripts-Creation - Target":"test_scripts_creation_target",
            "Test-scripts-Updation - Target":"test_scripts_updation_target",	
            "Test-scripts-Execution - Target":"test_scripts_execution_target",	
            "Project Documentation - Target":"project_doc_target",	
            "Internal Review - Target":	"internal_Review_target",
            "Regression Cycle - Target":"regression_cycle_target",	
            "Requirement analyzing/writing testcondition - Target":"req_anal_target",
            "End-End test cases executed - Target":	"end_cases_exec_target",
            "Site Scrub - Target"	:"site_Scrub_target",
            "Task Achivement/Coverage score - Target":	"task_coverage_score_target",
            "Assessment Test score - Target":"assessment_score_target",
            "Assessment Retest score - Target"	:"assessment_re_score_target",
            "Certification Test score - Target"	:"cert_score_target",
            "Certification Retest score - Target":"cert_re_score_target",	
            "New Features Implemented - Target":"new_features_imp_target",	
            "Defects Fixed-Target":	"defects_fixed_target",
            "Enhancements-Target":"enhancements_target",	
            "Figma Designs-Created-Target":	"fig_desgns_target",
            "Project Documentation Update-Target":"doc_update_target",
            "Research-Target":"research_target",	
            "Reporting Person":"reporting_manager"
            # Add more mappings as per your Excel file
            }
            mapped_columns = {column_mapping[h]: idx for idx, h in enumerate(headers) if h in column_mapping}
            processed_employees = set()
            # Process the rows and add to the database
            for row in ws.iter_rows(min_row=2, values_only=True):
                employee_data = {db_field: row[idx] for db_field, idx in mapped_columns.items()}
                if 'emp_name' in employee_data:
                    emp_name = employee_data['emp_name']
                    if emp_name is None:
                        name_exists = False 
                    else:
                        # Check if employee already exists
                        name_exists = db.session.query(
                            db.session.query(Employee_information)
                            .filter(func.lower(Employee_information.emp_name) == emp_name.lower())
                            .exists()
                        ).scalar()

                    # If employee already exists, flash the message
                    if name_exists:
                        if emp_name not in processed_employees:
                            flash(f"Employee with name '{emp_name}' already exists. Skipping entry.", "danger")
                            processed_employees.add(emp_name)  # Add to the set of processed employees
                        continue  # Skip this entry since it's a duplicate
                if 'emp_date' in employee_data:
                    emp_date = employee_data['emp_date']

                    if isinstance(emp_date, datetime):
                        emp_date = emp_date.date()  # Removes the time and keeps only the date
                        print("Date without time:", emp_date)

                    # If emp_date is a string, convert it to a date object
                    elif isinstance(emp_date, str):
                        try:
                            emp_date = datetime.strptime(emp_date, '%Y-%m-%d').date()  # Convert to date
                            print("Successfully parsed the date:", emp_date)
                        except ValueError:
                            print("Incorrect date format")
                            emp_date = None 
                    if emp_date:
                        try:
                            db.session.query(Employee_information).update({"emp_date": emp_date})
                            db.session.commit()
                            flash(f"Updated emp_date for all employees to {emp_date}.", "success")
                        except Exception as e:
                            print(e)
                            flash("Failed to update emp_date for existing employees.", "danger")          
                            
                if all(value is None for value in employee_data.values()):
                    print("Skipping row with all None values")
                    continue             
                
                
            
                employee = Employee_information(**employee_data)

                # Add the employee record to the session
                db.session.add(employee)
            
            # # Commit all changes to the database
            db.session.commit() 
            
        except Exception as e:
            print(e)   
            flash("Failed to upload employees. Please check the sample file.", "danger")
    return render_template("employee_upload.html")

@app.route("/dmax_table", methods=["GET", "POST"])
def dmax_table():
    default_page_size = 10
    page_size_options = [10, 20, 30, 'All']
    if request.method == "POST":
        selected_page_size = request.form.get("page_size", default_page_size)
        # Set page_size to None for 'All', or convert to int if numeric
        if selected_page_size == 'All':
            session['page_size'] = None
        else:
            session['page_size'] = int(selected_page_size)
        # Redirect to page 1 with the current search term to avoid form resubmission
        return redirect(url_for('dmax_table', page=1, search_term=request.args.get('search_term', '')))
    designation_counts = (
        db.session.query(Employee_information.emp_designation, func.count(Employee_information.emp_designation))
        .group_by(Employee_information.emp_designation)
        .all()
    )
    labels = [row[0] for row in designation_counts]  # Designation names
    scores = [row[1] for row in designation_counts]  # Counts
    page_size = session.get('page_size', default_page_size)
    search_term = request.args.get("search_term", "").strip()
    selected_designation=request.args.get("designation")
    selected_project=request.args.get("project")
    selected_month = request.args.get("month", default=str(datetime.now().month - 1 if datetime.now().month > 1 else 12))
    if not selected_month:
        current_month = datetime.now().month
        selected_month = str(current_month - 1 if current_month > 1 else 12)
    page = request.args.get("page", 1, type=int)
    projects=["Akyrian","Auxo","Avanti","Bench","Fora Travels","Indihood","IPS","IQHive","LevelBlue","Web Development","Opus Clip","Training"]
    designations=["Intern","Jr.QA Engineer","QA Engineer","Sr.QA Engineer","QA Lead"]
    query=Dform.query
    if search_term:
        query = query.filter(func.lower(Dform.employee_name) == search_term.lower())
    if selected_designation:
        query=query.filter(Dform.designation == selected_designation)
    if selected_project:
        query = query.filter(func.lower(Dform.project) == selected_project.strip().lower())
    if selected_month:
        query=query.filter(extract('month', Dform.today_date) == int(selected_month))  
    
    if page_size:  # If 'All' is not selected, paginate based on page size
        paginated_entries = query.paginate(page=page, per_page=page_size, error_out=False)
    else:
        # Show all results if page size is 'All'
        paginated_entries = query.paginate(page=1, per_page=query.count(), error_out=False)

    # Prepare a list to hold the data for all entries
    data_list = []
    
    # Check if there are any entries
    if paginated_entries.items:
        # Use SQLAlchemy's metadata to get the columns in the order they are defined
        columns =['employee_name','employee_id','project','designation','production', 'quality', 'attendance', 'skill','Dmax_score']  # Use the first entry to get column names
        
        # Loop through each entry to create a dictionary for each row
        for entry in paginated_entries.items:
            data = {field: getattr(entry, field) for field in columns}  # Create a dict for each entry
            data_list.append(data)  # Add to the list

    top_scores = query.order_by(Dform.Dmax_score.desc()).limit(4).all()
    chart_data = [{"name": entry.employee_name, "score": entry.Dmax_score} for entry in top_scores]
    

    return render_template('dmax_table.html', data_list=data_list,pagination=paginated_entries, search_term=search_term,page_size=page_size,page_size_options=page_size_options,chart_data=chart_data,selected_month=selected_month,selected_designation=selected_designation,selected_project=selected_project,projects=projects,designations=designations,labels=labels, scores=scores)

@app.route("/team_dmax_table", methods=["GET", "POST"])
def team_dmax_table():
    user_details = get_logged_in_user_details()
    if user_details:
        user_name=user_details['name']
        role=user_details['role']
        if role=="admin" or role=="super_admin":
            employees_under_manager = Employee_information.query.all()
        if role=="manager":    
            employees_under_manager = Employee_information.query.filter_by(reporting_manager=user_name).all()
        filtered_employees = []
        for emp in employees_under_manager:
            matched_employee = Employee_information.query.filter_by(emp_email=emp.emp_email).first()
            if matched_employee:
                filtered_employees.append({
                    "name": matched_employee.emp_name,
                    "project":matched_employee.emp_project,
                    "designation":matched_employee.emp_designation,
                    "date":matched_employee.emp_date,
                    "emp_id":matched_employee.emp_id,
                    "reporting_manager":matched_employee.reporting_manager,
                    "id":matched_employee.id
                })
                
        return render_template('team_dmax_table.html',employees=filtered_employees,user_name=user_name)
    
    return "No user found or not logged in."  

@app.route("/view_dscore", methods=["GET", "POST"])
def view_dscore():
    user_name = get_logged_in_user_details()
    ALLOWED_COLUMNS = [
        "id","employee_name","target", "actual", "production",
        "quality", "attendance", "skill", "new_initiatives", "Dmax_score"
    ]
    # ALLOWED_COLUMNS = [
    #     "employee_name", "today_date", "test_case_creation_target",
    #     "test_case_creation_actual", "test_case_updation_target", "test_case_updation_actual",
    #     "test_case_execution_target", "test_case_execution_actual", "defects_found_target",
    #     "defects_found_actual","defects_verification_target", "defects_verification_actual", "test_scripts_creation_target", "test_scripts_creation_actual",
    #     "test_scripts_execution_target","test_scripts_execution_actual","test_scripts_updation_target", "test_scripts_updation_actual",
    #     "site_Scrub_target", "site_Scrub_actual", "project_doc_target",
    #     "project_doc_actual", "internal_Review_target", "internal_Review_actual", "regression_cycle_target",
    #     "regression_cycle_actual", "req_anal_target", "req_anal_actual", "end_cases_exec_target",
    #     "end_cases_exec_actual", "task_coverage_score_target", "task_coverage_score_actual",
    #     "assessment_score_target", "assessment_score_actual", "assessment_re_score_target",
    #     "assessment_re_score_actual", "cert_score_target", "cert_score_actual", "cert_re_score_target",
    #     "cert_re_score_actual", "new_features_imp_target", "new_features_imp_actual", "defects_fixed_target",
    #     "defects_fixed_actual", "enhancements_target", "enhancements_actual", "fig_desgns_target",
    #     "fig_desgns_actual", "doc_update_target", "doc_update_actual", "research_target", "research_actual",
    #     "inv_defs", "spel_errors", "client_esc", "tst_cases_missing", "att", "dtouch", "new_init", "target", "actual", "production",
    #     "quality", "attendance", "skill", "new_initiatives", "Dmax_score"
    # ]
    
    if user_name:
        
        role = user_name['role']
        # email=user_name['email']
        user_name=user_name['name'].lower()
        search_query = request.args.get('search', '').strip().lower()
        selected_date = request.args.get('date') 
        selected_month = request.args.get('month')
        
        if role =="manager":    
            employees_under_manager = Employee_information.query.filter(func.lower(Employee_information.reporting_manager)==user_name).all()
            filtered_employees=[]
            for emp in employees_under_manager:
                matched_employees = get_filtered_employees(
                    Dform.query.filter_by(employee_email=emp.emp_email),
                    search_query,
                    selected_month,
                    selected_date
                )
                # query = Dform.query.filter_by(employee_email=emp.emp_email)
                # if search_query:
                #     # Use = for exact match (case-sensitive)
                #     query = query.filter(func.lower(Dform.employee_name) == search_query)
                    
                # if selected_month:
                #     # Extract the month from today_date (assuming today_date is a datetime field)
                #     # please import extract in godaddy
                #    query = query.filter(extract('month', Dform.today_date) == int(selected_month))
                # if selected_date:
                #     query = query.filter_by(today_date=selected_date)
                # matched_employees = query.all()    
                if matched_employees:
                    for matched in matched_employees:
                        filtered_employees.append(
                            {
                                column: getattr(matched, column, None)  # Use getattr to get the attribute dynamically
                                for column in ALLOWED_COLUMNS         # Filter by allowed columns
                            }
                        )
            return render_template("view_dscore.html",employees=filtered_employees,role=role,search_query=search_query, selected_month=selected_month, selected_date=selected_date)        

        if role=="crewmate":
             
            matched_employees = get_filtered_employees(
                Dform.query.filter_by(employee_email=email),
                search_query,
                selected_month,
                selected_date
            )
            filtered_employees=[]
            if matched_employees:
                for matched in matched_employees:
                    filtered_employees.append(
                            {
                                column: getattr(matched, column, None)  # Use getattr to get the attribute dynamically
                                for column in ALLOWED_COLUMNS         # Filter by allowed columns
                            }
                        )
                    

            return render_template("view_dscore.html",employees=filtered_employees,role=role,search_query=search_query, selected_month=selected_month, selected_date=selected_date)                
                
        if role == "admin" or role == "super_admin":
            filtered_employees = []
            matched_employees = get_filtered_employees(
                Dform.query,
                search_query,
                selected_month,
                selected_date
            )
            if matched_employees:
                for matched in matched_employees:
                    filtered_employees.append(
                            {
                                column: getattr(matched, column, None)  # Use getattr to get the attribute dynamically
                                for column in ALLOWED_COLUMNS         # Filter by allowed columns
                            }
                        )
            return render_template("view_dscore.html", employees=filtered_employees, role=role, search_query=search_query, selected_month=selected_month, selected_date=selected_date)

@app.route('/delete_employee/<int:id>', methods=['POST'])
def delete_employee(id):
    employee = Employee_information.query.get(id)
    if employee:
        db.session.delete(employee)
        db.session.commit()
        flash('Employee deleted successfully!', 'success')
    else:
        flash('Employee not found.', 'danger')
    return jsonify({'success': True})

@app.route('/operational_excellence/<string:emp_id>', methods=['GET', 'POST'])
def operational_excellence(emp_id):
    employee_info = Employee_information.query.filter_by(emp_id=emp_id).first()
    op_excellence=OperationalExcellence.query.filter_by(emp_id=emp_id).first()
    if not op_excellence:
        op_excellence = OperationalExcellence(
            emp_id=emp_id,
            attendance_score=0,
            dtouch_score=0,
            new_init_score=0
        )
        db.session.add(op_excellence)
        db.session.commit()
    
    if employee_info:
        designation = employee_info.emp_designation
        if request.method == 'POST':
            attendance=request.form.get('attendance')
            dtouch = request.form.get('dtouch')
            new_init=request.form.get('newInitiatives')
            attendance =int(attendance)
            dtouch=int(dtouch)   
            new_init=int(new_init) 
            if designation == "Intern":
                attendance = int((attendance * 10 / 100) * 100)
                dtouch = int(((attendance * 10 / 100 / 100) * 100) * 100)
                new_init=0
            if designation=="Jr.QA Engineer":
                attendance = int((attendance * 5 / 100) * 100)
                dtouch = int(((attendance * 10 / 100 / 100) * 100) * 100)
                new_init=0
            if designation=="QA Engineer":
                attendance = int((attendance * 5 / 100) * 100)
                dtouch = int(((attendance * 5 / 100 / 100) * 100) * 100)
                new_init = int(((new_init * 15 / 100 / 100) * 100) * 100)
            if designation=="Sr.QA Engineer":
                attendance = int((attendance * 5 / 100) * 100)
                dtouch = (((dtouch * 5 / 100 / 100) * 100) )
                new_init = ((new_init * 20 / 100 / 100) * 100)
                
            if designation=="QA Lead":
                attendance = int((attendance * 5 / 100) * 100)
                dtouch = (((dtouch * 5 / 100 / 100) * 100) )    
                new_init = ((new_init * 20 / 100 / 100) * 100)
            op_excellence.attendance_score = attendance
            op_excellence.dtouch_score = dtouch
            op_excellence.new_init_score = new_init    
            
            db.session.commit()    

            
    return render_template('operational_excellence.html',emp_id=emp_id)

@app.route("/full_table_view/<int:id>")
def full_table_view(id):
    employee = Dform.query.filter_by(id=id).first()
    ALLOWED_COLUMNS = [
        "employee_name", "today_date", "test_case_creation_target",
        "test_case_creation_actual", "test_case_updation_target", "test_case_updation_actual",
        "test_case_execution_target", "test_case_execution_actual", "defects_found_target",
        "defects_found_actual","defects_verification_target", "defects_verification_actual", "test_scripts_creation_target", "test_scripts_creation_actual",
        "test_scripts_execution_target","test_scripts_execution_actual","test_scripts_updation_target", "test_scripts_updation_actual",
        # # "site_Scrub_target", "site_Scrub_actual", "project_doc_target",
        # # "project_doc_actual", "internal_Review_target", "internal_Review_actual", "regression_cycle_target",
        # # "regression_cycle_actual", "req_anal_target", "req_anal_actual", "end_cases_exec_target",
        # # "end_cases_exec_actual", "task_coverage_score_target", "task_coverage_score_actual",
        # # "assessment_score_target", "assessment_score_actual", "assessment_re_score_target",
        # # "assessment_re_score_actual", "cert_score_target", "cert_score_actual", "cert_re_score_target",
        # # "cert_re_score_actual", "new_features_imp_target", "new_features_imp_actual", "defects_fixed_target",
        # # "defects_fixed_actual", "enhancements_target", "enhancements_actual", "fig_desgns_target",
        # # "fig_desgns_actual", "doc_update_target", "doc_update_actual", "research_target", "research_actual",
        "inv_defs",  "spel_errors",  "client_esc", "tst_cases_missing", "attendance","target","actual","production","quality","attendance","skill","new_initiatives","Dmax_score"
        # "quality", "attendance", "skill", "new_initiatives", "Dmax_score"
    ]
    if employee:
        return render_template("full_table_view.html", employee=employee, ALLOWED_COLUMNS=ALLOWED_COLUMNS)
with app.app_context():
        
        db.create_all()
        

if __name__ == "__main__":
    app.run(debug=True)